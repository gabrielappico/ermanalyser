"""Extract ESG answers from the Itaú Unibanco 2024 .xlsm spreadsheet."""

import openpyxl
import json
import re

XLSX_PATH = r"c:\Users\gabri\OneDrive\Área de Trabalho\ERM\Itaú Unibanco Holding S.A. 2024.xlsm"

# These are the ESG theme sheets (skip meta/admin sheets)
ESG_SHEETS = [
    "Controvérsias", "Materialidade",
    "Recursos Hídricos", "Recursos Energéticos",
    "Sistema Gestão Socioambiental", "Materiais básicos",
    "Resíduos Sólidos", "Efluentes líquidos",
    "Emissões atmosféricas", "Mudanças Climáticas Mitigação",
    "Mudanças Climáticas Adaptação", "Ecossistemas",
    "Saúde e segurança", "Condições de trabalho",
    "Gestão de carreira", "Diversidade",
    "Segurança de dados", "Qualidade e segurança produtos",
    "Ecodesign", "Desenvolvimento",
    "Direitos Humanos", "Temas sociais na cadeia",
    "Temas ambientais na cadeia",
    "Remuneração executivos", "Conselho e Diretoria",
    "Minoritários", "Integridade", "Transparência",
    "Resultado"
]

def explore_all_sheets(path: str):
    """Dump first 50 rows of every ESG sheet to understand structure."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    
    print(f"Total sheets: {len(wb.sheetnames)}")
    print(f"All sheets: {wb.sheetnames}\n")
    
    all_data = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'='*70}")
        print(f"SHEET: '{sheet_name}'")
        print(f"{'='*70}")
        
        rows_data = []
        for row_idx, row in enumerate(ws.iter_rows(max_row=50, values_only=False), 1):
            cells = {}
            for cell in row:
                if cell.value is not None:
                    col_letter = openpyxl.utils.get_column_letter(cell.column)
                    cells[col_letter] = str(cell.value)[:120]
            
            if cells:
                rows_data.append({"row": row_idx, "cells": cells})
                print(f"  Row {row_idx:3d}: {cells}")
        
        all_data[sheet_name] = rows_data
    
    wb.close()
    return all_data


def extract_all_answers(path: str):
    """Extract answers from ALL ESG sheets with flexible column detection."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    
    results = {}
    total_questions = 0
    total_answered = 0
    
    for sheet_name in wb.sheetnames:
        if sheet_name in ["Capa", "Sumário", "Conceito&Orientações", "Dados da empresa", 
                          "Abrev_Temas", "Assinaturas", "Padronização_dados", "Padronização", "Notas"]:
            continue
            
        ws = wb[sheet_name]
        sheet_answers = []
        
        for row_idx, row in enumerate(ws.iter_rows(values_only=False), 1):
            cells = {}
            for cell in row:
                if cell.value is not None:
                    try:
                        col_letter = openpyxl.utils.get_column_letter(cell.column)
                        cells[col_letter] = cell.value
                    except (AttributeError, TypeError):
                        pass
            
            row_num = row_idx
            
            # Try to find question ID pattern (X.Y.Z)
            question_id = None
            question_text = None
            answer = None
            source = None
            
            # Check columns B, C, D for question ID
            for col in ["B", "C", "D", "A"]:
                val = cells.get(col)
                if val is not None:
                    s = str(val).strip()
                    match = re.match(r'^(\d+\.\d+(?:\.\d+)*(?:\.\d+)*)', s)
                    if match:
                        question_id = match.group(1)
                        # The question text might be in the same cell after the ID
                        rest = s[len(question_id):].strip()
                        if rest:
                            question_text = rest
                        break
            
            # Look for question text in C, D columns
            if not question_text:
                for col in ["C", "D"]:
                    val = cells.get(col)
                    if val and isinstance(val, str) and len(val) > 10:
                        question_text = val
                        break
            
            # Answer is typically in column E
            answer = cells.get("E")
            
            # Source/evidence in column F or L
            source = cells.get("F") or cells.get("L")
            
            if question_id or (question_text and len(str(question_text or "")) > 15):
                entry = {
                    "row": row_num,
                    "question_id": question_id,
                    "question_text": str(question_text)[:500] if question_text else None,
                    "answer": str(answer).strip() if answer is not None else None,
                    "source": str(source)[:1000] if source else None,
                }
                sheet_answers.append(entry)
                total_questions += 1
                if answer is not None and str(answer).strip() not in ("", "None", "N/A"):
                    total_answered += 1
        
        if sheet_answers:
            results[sheet_name] = sheet_answers
    
    wb.close()
    
    print(f"\n{'='*60}")
    print(f"EXTRACTION SUMMARY")
    print(f"{'='*60}")
    print(f"Sheets with data: {len(results)}")
    print(f"Total questions found: {total_questions}")
    print(f"Questions with answers: {total_answered}")
    print(f"\nPer-sheet breakdown:")
    for sheet, data in results.items():
        answered = sum(1 for d in data if d["answer"] and d["answer"] not in ("None", "", "N/A"))
        print(f"  {sheet:40s}: {len(data):3d} questions, {answered:3d} answered")
    
    return results


if __name__ == "__main__":
    import sys
    mode = sys.argv[1] if len(sys.argv) > 1 else "extract"
    
    if mode == "explore":
        data = explore_all_sheets(XLSX_PATH)
    else:
        answers = extract_all_answers(XLSX_PATH)
        output_path = XLSX_PATH.replace(".xlsm", "_answers.json")
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(answers, f, ensure_ascii=False, indent=2)
        print(f"\nSaved to: {output_path}")
