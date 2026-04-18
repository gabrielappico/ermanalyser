"""Live ESG Analysis Viewer — Terminal Rich UI.

Replays a completed ESG analysis from Supabase with animated terminal output,
showing each theme and question being filled in real-time.
Exports final results to Excel (.xlsx).

Usage:
    python live_viewer.py                   # auto-detect latest analysis
    python live_viewer.py --analysis-id ID  # specific analysis
    python live_viewer.py --fast            # faster replay (50ms delay)
    python live_viewer.py --no-excel        # skip Excel export
"""

import sys
import os
import io
import time
import argparse
from datetime import datetime

# Fix Windows encoding — force UTF-8 output
if sys.platform == "win32":
    os.system("")  # Enable ANSI escape codes on Windows
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.layout import Layout
from rich.live import Live
from rich.text import Text
from rich.progress import Progress, BarColumn, TextColumn, SpinnerColumn
from rich.columns import Columns
from rich.align import Align
from rich import box

from app.database import get_supabase

# ─── Config ──────────────────────────────────────────────────────────────────

DIMENSION_COLORS = {
    "environmental": "green",
    "social": "cyan",
    "governance": "yellow",
}

DIMENSION_EMOJI = {
    "environmental": "🌿",
    "social": "👥",
    "governance": "⚖️",
}

DIMENSION_LABELS = {
    "environmental": "AMBIENTAL",
    "social": "SOCIAL",
    "governance": "GOVERNANÇA",
}

RATING_COLORS = {
    "A": "bold green",
    "B": "blue",
    "C": "yellow",
    "D": "red",
    "E": "bold red",
}

ANSWER_ICONS = {
    "Sim": "✅",
    "Não": "❌",
    "N/A": "⬜",
}


def get_rating(score: float) -> str:
    if score >= 8.0:
        return "A"
    elif score >= 6.0:
        return "B"
    elif score >= 4.0:
        return "C"
    elif score >= 2.0:
        return "D"
    return "E"


# ─── Data Fetching ───────────────────────────────────────────────────────────

def fetch_analysis_data(analysis_id: str | None = None) -> dict:
    """Fetch all analysis data from Supabase, organized by theme."""
    sb = get_supabase()

    if analysis_id:
        analysis = sb.table("analyses").select(
            "*, companies(name, ticker, sector)"
        ).eq("id", analysis_id).single().execute().data
    else:
        analysis = sb.table("analyses").select(
            "*, companies(name, ticker, sector)"
        ).eq("status", "completed").order("completed_at", desc=True).limit(1).single().execute().data

    if not analysis:
        raise ValueError("No completed analysis found.")

    aid = analysis["id"]

    themes = sb.table("esg_themes").select("*").order("display_order").execute().data

    questions = sb.table("esg_questions").select("*").order("display_order").execute().data

    answers = sb.table("answers").select("*").eq("analysis_id", aid).execute().data

    theme_scores = sb.table("theme_scores").select("*").eq("analysis_id", aid).execute().data

    # Build lookup maps
    answer_map = {a["question_id"]: a for a in answers}
    score_map = {ts["theme_id"]: ts for ts in theme_scores}
    theme_questions = {}
    for q in questions:
        tid = q["theme_id"]
        if tid not in theme_questions:
            theme_questions[tid] = []
        theme_questions[tid].append(q)

    return {
        "analysis": analysis,
        "themes": themes,
        "theme_questions": theme_questions,
        "answer_map": answer_map,
        "score_map": score_map,
    }


# ─── Display Components ─────────────────────────────────────────────────────

def build_header(analysis: dict) -> Panel:
    """Build the header panel with company info."""
    company = analysis.get("companies", {})
    name = company.get("name", "N/A")
    sector = company.get("sector", "N/A")
    year = analysis.get("report_year", "N/A")
    ticker = company.get("ticker", "")

    title_text = Text()
    title_text.append("  📊 ESG LIVE ANALYZER", style="bold white")
    title_text.append(f"\n\n  Empresa: ", style="dim")
    title_text.append(f"{name}", style="bold white")
    if ticker:
        title_text.append(f" ({ticker})", style="dim cyan")
    title_text.append(f"\n  Setor:   ", style="dim")
    title_text.append(f"{sector}", style="white")
    title_text.append(f"\n  Ano:     ", style="dim")
    title_text.append(f"{year}", style="white")

    return Panel(
        title_text,
        border_style="bright_blue",
        box=box.DOUBLE_EDGE,
        padding=(1, 2),
    )


def build_dimension_bars(dim_progress: dict, dim_totals: dict) -> Panel:
    """Build the dimension progress bars."""
    lines = []
    for dim in ["environmental", "social", "governance"]:
        emoji = DIMENSION_EMOJI[dim]
        label = DIMENSION_LABELS[dim]
        color = DIMENSION_COLORS[dim]
        done = dim_progress.get(dim, 0)
        total = dim_totals.get(dim, 0)

        if total == 0:
            bar_str = "░" * 30
            pct_str = " --"
        else:
            pct = done / total
            filled = int(pct * 30)
            bar_str = "█" * filled + "░" * (30 - filled)
            pct_str = f"{pct * 100:3.0f}%"

        line = Text()
        line.append(f"  {emoji} {label:12s} ", style=f"bold {color}")
        line.append(f"[{bar_str}]", style=color)
        line.append(f" {pct_str}", style=f"bold {color}")
        line.append(f"  ({done}/{total})", style="dim")
        lines.append(line)

    content = Text("\n").join(lines)
    return Panel(content, title="📈 Progresso por Dimensão", border_style="blue", padding=(1, 1))


def build_score_panel(dim_scores: dict, overall_score: float | None) -> Panel:
    """Build final scores panel."""
    lines = []
    for dim in ["environmental", "social", "governance"]:
        emoji = DIMENSION_EMOJI[dim]
        label = DIMENSION_LABELS[dim]
        color = DIMENSION_COLORS[dim]
        score = dim_scores.get(dim)

        line = Text()
        line.append(f"  {emoji} {label:12s}  ", style=f"bold {color}")
        if score is not None:
            rating = get_rating(score)
            line.append(f"{score:5.2f}", style=f"bold {color}")
            line.append(f"  ({rating})", style=RATING_COLORS.get(rating, "white"))
        else:
            line.append("  --  ", style="dim")
        lines.append(line)

    if overall_score is not None:
        lines.append(Text())
        overall_line = Text()
        overall_line.append("  🏆 OVERALL       ", style="bold white")
        rating = get_rating(overall_score)
        overall_line.append(f"{overall_score:5.2f}", style="bold white")
        overall_line.append(f"  ({rating})", style=RATING_COLORS.get(rating, "white"))
        lines.append(overall_line)

    content = Text("\n").join(lines)
    return Panel(content, title="🎯 Scores Finais", border_style="bright_green", padding=(1, 1))


def build_theme_header(theme: dict, theme_score: dict | None) -> Text:
    """Build a single theme header line."""
    dim = theme["dimension"]
    color = DIMENSION_COLORS[dim]
    emoji = DIMENSION_EMOJI[dim]
    name = theme["name"]
    number = theme.get("theme_number", "?")

    line = Text()
    line.append(f"\n  {emoji} ", style=color)
    line.append(f"Tema {number}: {name}", style=f"bold {color}")

    if theme_score:
        score = theme_score.get("raw_score", 0)
        rating = theme_score.get("rating", "?")
        line.append(f"  →  {score:.2f} ({rating})", style=RATING_COLORS.get(rating, "white"))

    return line


def build_question_line(q: dict, answer: dict | None, show_answer: bool = True) -> Text:
    """Build a single question line."""
    qid = q.get("question_id", "?")
    text = q.get("question_text", "")

    # Clean up question text — remove question_id prefix if present
    if text.startswith(qid):
        text = text[len(qid):].lstrip(" .")

    # Truncate long questions
    max_len = 70
    if len(text) > max_len:
        text = text[:max_len - 3] + "..."

    line = Text()

    if show_answer and answer:
        ans_val = answer.get("answer", "N/A")
        icon = ANSWER_ICONS.get(ans_val, "❓")
        conf = answer.get("confidence_score", 0)

        line.append(f"    {icon} ", style="white")
        line.append(f"{qid:8s} ", style="dim cyan")
        line.append(f"{text:72s} ", style="white")
        line.append(f"{ans_val:4s}", style="bold green" if ans_val == "Sim" else ("bold red" if ans_val == "Não" else "dim"))
        line.append(f"  {conf:.2f}", style="dim yellow")
    else:
        line.append(f"    ⏳ ", style="dim")
        line.append(f"{qid:8s} ", style="dim")
        line.append(f"{text:72s} ", style="dim white")
        line.append(f" ...", style="dim")

    return line


# ─── Live Replay ─────────────────────────────────────────────────────────────

def run_replay(data: dict, delay: float = 0.12, console: Console | None = None):
    """Run the animated replay of the analysis."""
    if console is None:
        console = Console()

    analysis = data["analysis"]
    themes = data["themes"]
    theme_questions = data["theme_questions"]
    answer_map = data["answer_map"]
    score_map = data["score_map"]

    # Calculate dimension totals
    dim_totals = {"environmental": 0, "social": 0, "governance": 0}
    dim_progress = {"environmental": 0, "social": 0, "governance": 0}
    dim_scores_running = {"environmental": None, "social": None, "governance": None}

    for theme in themes:
        qs = theme_questions.get(theme["id"], [])
        dim_totals[theme["dimension"]] += len(qs)

    # Print header (static)
    console.clear()
    console.print(build_header(analysis))
    console.print()

    total_questions = sum(dim_totals.values())
    answered_count = 0

    # Main progress + spinner
    with Progress(
        SpinnerColumn("dots"),
        TextColumn("[bold blue]{task.description}"),
        BarColumn(bar_width=40, complete_style="green", finished_style="bright_green"),
        TextColumn("[bold]{task.percentage:>3.0f}%"),
        TextColumn("({task.completed}/{task.total})"),
        console=console,
        transient=False,
    ) as progress:
        main_task = progress.add_task("Análise ESG", total=total_questions)

        current_dim = None

        for theme in themes:
            theme_id = theme["id"]
            dimension = theme["dimension"]
            theme_name = theme["name"]
            theme_number = theme.get("theme_number", "?")
            qs = theme_questions.get(theme_id, [])

            if not qs:
                continue

            # Print dimension separator when dimension changes
            if dimension != current_dim:
                current_dim = dimension
                color = DIMENSION_COLORS[dimension]
                emoji = DIMENSION_EMOJI[dimension]
                label = DIMENSION_LABELS[dimension]
                console.print()
                console.rule(f"{emoji}  {label}", style=color)
                console.print()
                time.sleep(delay * 3)

            # Theme header (show as "analyzing...")
            ts = score_map.get(theme_id)
            header_text = Text()
            header_text.append(f"  📋 Tema {theme_number}: ", style="bold white")
            header_text.append(f"{theme_name}", style=f"bold {DIMENSION_COLORS[dimension]}")
            header_text.append(f"  ({len(qs)} perguntas)", style="dim")
            console.print(header_text)
            time.sleep(delay * 2)

            # Animate each question
            for q in qs:
                answer = answer_map.get(q["id"])
                line = build_question_line(q, answer, show_answer=True)
                console.print(line)

                answered_count += 1
                dim_progress[dimension] += 1
                progress.update(main_task, completed=answered_count)
                time.sleep(delay)

            # Show theme score
            if ts:
                score = ts.get("raw_score", 0)
                rating = ts.get("rating", "?")
                score_line = Text()
                score_line.append(f"    ── Score: ", style="dim")
                score_line.append(f"{score:.2f}", style=f"bold {RATING_COLORS.get(rating, 'white')}")
                score_line.append(f" ({rating})", style=RATING_COLORS.get(rating, "white"))
                score_line.append(f" ──", style="dim")
                console.print(score_line)
            console.print()
            time.sleep(delay * 2)

    # Final scores
    console.print()
    console.print(build_dimension_bars(dim_progress, dim_totals))
    console.print()

    dim_final = {
        "environmental": analysis.get("environmental_score"),
        "social": analysis.get("social_score"),
        "governance": analysis.get("governance_score"),
    }
    console.print(build_score_panel(dim_final, analysis.get("overall_score")))

    return data


# ─── Excel Export ────────────────────────────────────────────────────────────

def export_to_excel(data: dict, output_path: str | None = None):
    """Export analysis results to a formatted Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    analysis = data["analysis"]
    themes = data["themes"]
    theme_questions = data["theme_questions"]
    answer_map = data["answer_map"]
    score_map = data["score_map"]

    company = analysis.get("companies", {})
    company_name = company.get("name", "Empresa")
    year = analysis.get("report_year", datetime.now().year)

    if output_path is None:
        output_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "..",
            f"{company_name} {year} - ESG Rating.xlsx"
        )

    wb = Workbook()

    # ── Styles ──
    header_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    dim_fonts = {
        "environmental": Font(name="Calibri", size=12, bold=True, color="2E7D32"),
        "social": Font(name="Calibri", size=12, bold=True, color="0277BD"),
        "governance": Font(name="Calibri", size=12, bold=True, color="F57F17"),
    }
    dim_fills = {
        "environmental": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
        "social": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
        "governance": PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),
    }
    theme_font = Font(name="Calibri", size=11, bold=True)
    q_font = Font(name="Calibri", size=10)
    ans_yes_font = Font(name="Calibri", size=10, bold=True, color="2E7D32")
    ans_no_font = Font(name="Calibri", size=10, bold=True, color="C62828")
    ans_na_font = Font(name="Calibri", size=10, color="9E9E9E")
    score_font = Font(name="Calibri", size=11, bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    # ── Summary Sheet ──
    ws_summary = wb.active
    ws_summary.title = "Resumo"
    ws_summary.sheet_properties.tabColor = "1F4E79"

    # Header
    ws_summary.merge_cells("A1:F1")
    ws_summary["A1"] = f"📊 ESG Rating — {company_name} ({year})"
    ws_summary["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws_summary["A1"].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ws_summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 40

    # Overall scores
    row = 3
    labels = [
        ("Overall Score", analysis.get("overall_score", 0), analysis.get("overall_rating", "?")),
        ("Environmental", analysis.get("environmental_score", 0), get_rating(analysis.get("environmental_score", 0))),
        ("Social", analysis.get("social_score", 0), get_rating(analysis.get("social_score", 0))),
        ("Governance", analysis.get("governance_score", 0), get_rating(analysis.get("governance_score", 0))),
    ]
    for label, score, rating in labels:
        ws_summary[f"A{row}"] = label
        ws_summary[f"A{row}"].font = Font(name="Calibri", size=12, bold=True)
        ws_summary[f"B{row}"] = round(score, 2) if score else 0
        ws_summary[f"B{row}"].font = score_font
        ws_summary[f"C{row}"] = rating
        ws_summary[f"C{row}"].font = Font(name="Calibri", size=12, bold=True,
                                          color="2E7D32" if rating == "A" else
                                                "0277BD" if rating == "B" else
                                                "F57F17" if rating == "C" else "C62828")
        row += 1

    # Theme scores table
    row += 2
    ws_summary[f"A{row}"] = "Tema"
    ws_summary[f"B{row}"] = "Dimensão"
    ws_summary[f"C{row}"] = "Score"
    ws_summary[f"D{row}"] = "Rating"
    ws_summary[f"E{row}"] = "Perguntas"
    for col in ["A", "B", "C", "D", "E"]:
        ws_summary[f"{col}{row}"].font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        ws_summary[f"{col}{row}"].fill = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
    row += 1

    for theme in themes:
        qs = theme_questions.get(theme["id"], [])
        ts = score_map.get(theme["id"])
        dim = theme["dimension"]

        ws_summary[f"A{row}"] = f"{theme.get('theme_number', '?')}. {theme['name']}"
        ws_summary[f"A{row}"].font = q_font
        ws_summary[f"B{row}"] = DIMENSION_LABELS.get(dim, dim)
        ws_summary[f"B{row}"].font = Font(name="Calibri", size=10, color=
                                          "2E7D32" if dim == "environmental" else
                                          "0277BD" if dim == "social" else "F57F17")
        ws_summary[f"C{row}"] = round(ts["raw_score"], 2) if ts and ts.get("raw_score") else 0
        ws_summary[f"D{row}"] = ts.get("rating", "-") if ts else "-"
        ws_summary[f"E{row}"] = len(qs)
        row += 1

    # Column widths for summary
    ws_summary.column_dimensions["A"].width = 40
    ws_summary.column_dimensions["B"].width = 15
    ws_summary.column_dimensions["C"].width = 10
    ws_summary.column_dimensions["D"].width = 8
    ws_summary.column_dimensions["E"].width = 12

    # ── Detail Sheet (per dimension) ──
    for dim in ["environmental", "social", "governance"]:
        dim_label = DIMENSION_LABELS[dim]
        ws = wb.create_sheet(title=dim_label[:31])
        ws.sheet_properties.tabColor = (
            "4CAF50" if dim == "environmental" else
            "2196F3" if dim == "social" else "FFC107"
        )

        # Header
        headers = ["ID", "Pergunta", "Resposta", "Confiança", "Justificativa", "Fonte", "Melhorias"]
        col_widths = [10, 60, 10, 12, 50, 30, 30]

        for c_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=c_idx, value=h)
            cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[chr(64 + c_idx)].width = w

        ws.row_dimensions[1].height = 25
        ws.auto_filter.ref = f"A1:G1"

        row = 2
        dim_themes = [t for t in themes if t["dimension"] == dim]
        for theme in dim_themes:
            qs = theme_questions.get(theme["id"], [])
            if not qs:
                continue

            ts = score_map.get(theme["id"])

            # Theme separator row
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            theme_cell = ws.cell(row=row, column=1)
            theme_score_str = f" — {ts['raw_score']:.2f} ({ts['rating']})" if ts else ""
            theme_cell.value = f"📋 Tema {theme.get('theme_number', '?')}: {theme['name']}{theme_score_str}"
            theme_cell.font = dim_fonts[dim]
            theme_cell.fill = dim_fills[dim]
            theme_cell.alignment = Alignment(vertical="center")
            ws.row_dimensions[row].height = 28
            row += 1

            for q in qs:
                answer = answer_map.get(q["id"])
                qid = q.get("question_id", "?")
                qtext = q.get("question_text", "")
                if qtext.startswith(qid):
                    qtext = qtext[len(qid):].lstrip(" .")

                ws.cell(row=row, column=1, value=qid).font = Font(name="Calibri", size=9, color="546E7A")
                ws.cell(row=row, column=2, value=qtext).font = q_font
                ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")

                if answer:
                    ans_val = answer.get("answer", "N/A")
                    ans_cell = ws.cell(row=row, column=3, value=ans_val)
                    ans_cell.font = ans_yes_font if ans_val == "Sim" else (ans_no_font if ans_val == "Não" else ans_na_font)
                    ans_cell.alignment = Alignment(horizontal="center")

                    ws.cell(row=row, column=4, value=round(answer.get("confidence_score", 0), 2)).font = q_font
                    ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")

                    ws.cell(row=row, column=5, value=answer.get("justification", "")).font = Font(name="Calibri", size=9)
                    ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True, vertical="top")

                    ws.cell(row=row, column=6, value=answer.get("source_reference", "")).font = Font(name="Calibri", size=9)
                    ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True, vertical="top")

                    ws.cell(row=row, column=7, value=answer.get("improvement_points", "")).font = Font(name="Calibri", size=9)
                    ws.cell(row=row, column=7).alignment = Alignment(wrap_text=True, vertical="top")
                else:
                    ws.cell(row=row, column=3, value="—").font = ans_na_font

                # Apply borders
                for c in range(1, 8):
                    ws.cell(row=row, column=c).border = thin_border

                row += 1

            row += 1  # Blank row between themes

    wb.save(output_path)
    return output_path


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Live ESG Analysis Viewer")
    parser.add_argument("--analysis-id", type=str, help="Specific analysis ID to replay")
    parser.add_argument("--fast", action="store_true", help="Faster replay (50ms delay)")
    parser.add_argument("--slow", action="store_true", help="Slower replay (250ms delay)")
    parser.add_argument("--no-excel", action="store_true", help="Skip Excel export")
    parser.add_argument("--delay", type=float, default=None, help="Custom delay in seconds")
    args = parser.parse_args()

    delay = args.delay or (0.05 if args.fast else 0.25 if args.slow else 0.12)

    console = Console()

    # Intro animation
    console.clear()
    intro = Text()
    intro.append("\n\n")
    intro.append("  ╔══════════════════════════════════════════════════╗\n", style="bright_blue")
    intro.append("  ║                                                  ║\n", style="bright_blue")
    intro.append("  ║", style="bright_blue")
    intro.append("   📊 ESG LIVE ANALYZER — Replay Mode  ", style="bold white")
    intro.append("         ║\n", style="bright_blue")
    intro.append("  ║                                                  ║\n", style="bright_blue")
    intro.append("  ╚══════════════════════════════════════════════════╝\n", style="bright_blue")
    console.print(intro)
    time.sleep(1)

    # Fetch data
    console.print("  🔌 Conectando ao Supabase...", style="dim")
    time.sleep(0.5)

    try:
        data = fetch_analysis_data(args.analysis_id)
    except Exception as e:
        console.print(f"\n  ❌ Erro: {e}", style="bold red")
        return 1

    analysis = data["analysis"]
    company = analysis.get("companies", {})
    console.print(f"  ✅ Análise encontrada: [bold]{company.get('name', '?')}[/] ({analysis.get('report_year', '?')})")
    console.print(f"  📋 {len(data['themes'])} temas, {len(data['answer_map'])} respostas")
    console.print()
    time.sleep(1)

    console.print("  ▶  Iniciando replay...\n", style="bold green")
    time.sleep(0.8)

    # Run replay
    run_replay(data, delay=delay, console=console)

    # Excel export
    if not args.no_excel:
        console.print()
        console.print("  📄 Exportando para Excel...", style="dim")
        try:
            path = export_to_excel(data)
            console.print(f"  ✅ Excel salvo: [bold green]{path}[/]")
        except Exception as e:
            console.print(f"  ❌ Erro no Excel: {e}", style="red")

    console.print()
    console.print("  🏁 Replay finalizado!\n", style="bold bright_green")
    return 0


if __name__ == "__main__":
    sys.exit(main())
