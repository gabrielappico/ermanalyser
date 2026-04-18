"""Export ESG analysis results by filling the original ERM template (.xlsm).

Strategy: copy the template, open with openpyxl (preserving VBA macros),
fill in company data + answers, and save. The template's own formulas
auto-calculate scores/ratings when the file is opened in Excel.
"""

import os
import re
import shutil
import tempfile
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from app.database import get_supabase

# Regex to strip XML-illegal control characters that openpyxl rejects.
# Keeps tab (\x09), newline (\x0A), carriage return (\x0D).
_ILLEGAL_XML_CHARS = re.compile(
    r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F-\x9F]'
)


def _sanitize_for_excel(value: str | None) -> str | None:
    """Remove control characters that are illegal in Excel/XML cells."""
    if not value or not isinstance(value, str):
        return value
    return _ILLEGAL_XML_CHARS.sub('', value)


TEMPLATE_PATH = os.path.join(
    os.path.dirname(__file__), "..", "..", "..", "Rating ESG Modelo (1).xlsm"
)
TEMPLATE_PATH = os.path.normpath(TEMPLATE_PATH)

# Map DB theme names → spreadsheet sheet names
THEME_TO_SHEET = {
    "Recursos Hídricos": "Recursos Hídricos",
    "Recursos Energéticos": "Recursos Energéticos",
    "Sistema Gestão Socioambiental": "Sistema Gestão Socioambiental",
    "Materiais básicos": "Materiais básicos",
    "Resíduos Sólidos": "Resíduos Sólidos",
    "Efluentes líquidos": "Efluentes líquidos",
    "Emissões atmosféricas": "Emissões atmosféricas",
    "Mudanças Climáticas Mitigação": "Mudanças Climáticas Mitigação",
    "Mudanças Climáticas Adaptação": "Mudanças Climáticas Adaptação",
    "Ecossistemas": "Ecossistemas",
    "Saúde e segurança": "Saúde e segurança",
    "Condições de trabalho": "Condições de trabalho",
    "Gestão de carreira": "Gestão de carreira",
    "Diversidade": "Diversidade",
    "Segurança de dados": "Segurança de dados",
    "Qualidade e segurança produtos": "Qualidade e segurança produtos",
    "Ecodesign": "Ecodesign",
    "Desenvolvimento": "Desenvolvimento",
    "Direitos Humanos": "Direitos Humanos",
    "Temas sociais na cadeia": "Temas sociais na cadeia",
    "Temas ambientais na cadeia": "Temas ambientais na cadeia",
    "Remuneração executivos": "Remuneração executivos",
    "Conselho e Diretoria": "Conselho e Diretoria",
    "Minoritários": "Minoritários",
    "Integridade": "Integridade",
    "Transparência": "Transparência",
}

# Map common Brazilian sector names to GICS sectors used in the Materialidade sheet.
# The Peso formula references Capa!G35 and does MATCH against Materialidade row 7.
SECTOR_TO_GICS = {
    "Energia": "Energia",
    "Energy": "Energia",
    "Materiais": "Materiais",
    "Materials": "Materiais",
    "Papel e Celulose": "Materiais",
    "Celulose": "Materiais",
    "Mineracao": "Materiais",
    "Siderurgia": "Materiais",
    "Industrial": "Industrial",
    "Industrials": "Industrial",
    "Bens Industriais": "Industrial",
    "Consumo ciclico": "Consumo cíclico",
    "Consumer Discretionary": "Consumo cíclico",
    "Varejo": "Consumo cíclico",
    "Consumo nao ciclico": "Consumo não cíclico",
    "Consumer Staples": "Consumo não cíclico",
    "Alimentos": "Consumo não cíclico",
    "Bebidas": "Consumo não cíclico",
    "Saude": "Saúde",
    "Health Care": "Saúde",
    "Financeiro": "Serviços financeiros",
    "Financials": "Serviços financeiros",
    "Servicos financeiros": "Serviços financeiros",
    "Bancos": "Serviços financeiros",
    "Seguros": "Serviços financeiros",
    "Tecnologia": "Tecnologia da informação",
    "Information Technology": "Tecnologia da informação",
    "Comunicacoes": "Comunicações",
    "Telecomunicacoes": "Comunicações",
    "Communication Services": "Comunicações",
    "Utilities": "Utilities",
    "Utilidade publica": "Utilities",
    "Energia eletrica": "Utilities",
    "Saneamento": "Utilities",
    "Real estate": "Real estate",
    "Imoveis": "Real estate",
    "Construcao": "Real estate",
}


def _normalize_sector_to_gics(sector: str) -> str:
    """Map a company sector name to its GICS equivalent for the Materialidade sheet."""
    if not sector:
        return "Materiais"  # safe default
    # Direct match
    if sector in SECTOR_TO_GICS:
        return SECTOR_TO_GICS[sector]
    # Case-insensitive, accent-insensitive fuzzy match
    sector_lower = sector.lower().strip()
    for key, gics in SECTOR_TO_GICS.items():
        if key.lower() == sector_lower:
            return gics
    # Substring match
    for key, gics in SECTOR_TO_GICS.items():
        if key.lower() in sector_lower or sector_lower in key.lower():
            return gics
    return "Materiais"  # fallback


def _build_question_row_map(ws) -> dict[str, int]:
    """Scan a theme worksheet and build {question_id: row_number} mapping.

    Questions are identified by their ID pattern (e.g. '1.1.1', '13.2.1.3')
    found at the start of column B text.
    """
    qid_to_row = {}
    for row in ws.iter_rows(min_row=1, values_only=False):
        for cell in row:
            col = get_column_letter(cell.column)
            if col == "B" and cell.value:
                text = str(cell.value).strip()
                match = re.match(r"^(\d+\.\d+(?:\.\d+)*)", text)
                if match:
                    qid = match.group(1)
                    qid_to_row[qid] = cell.row
    return qid_to_row


def _find_company_name_cells(ws) -> list[tuple[int, int]]:
    """Find cells containing '[Razão Social da empresa]' placeholder."""
    cells = []
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value and "[Raz" in str(cell.value) and "empresa" in str(cell.value):
                cells.append((cell.row, cell.column))
    return cells


def _find_sector_cells(ws) -> list[tuple[int, int]]:
    """Find cells containing '[SETOR]' placeholder."""
    cells = []
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value and str(cell.value).strip() == "[SETOR]":
                cells.append((cell.row, cell.column))
    return cells


async def export_analysis_to_excel(analysis_id: str) -> str:
    """Generate an Excel file by filling the ERM template with analysis results.

    Returns the path to the generated .xlsm file.
    """
    sb = get_supabase()

    # Fetch all required data
    analysis = (
        sb.table("analyses")
        .select("*")
        .eq("id", analysis_id)
        .single()
        .execute()
        .data
    )
    company = (
        sb.table("companies")
        .select("*")
        .eq("id", analysis["company_id"])
        .single()
        .execute()
        .data
    )
    themes = (
        sb.table("esg_themes")
        .select("*")
        .order("display_order")
        .execute()
        .data
    )
    answers = (
        sb.table("answers")
        .select("*, esg_questions(question_id, question_text, section, theme_id)")
        .eq("analysis_id", analysis_id)
        .execute()
        .data
    )

    # Build lookup: theme_id → theme info
    theme_map = {t["id"]: t for t in themes}

    # Build lookup: (theme_name, question_id) → answer data
    answer_lookup: dict[tuple[str, str], dict] = {}
    for a in answers:
        q = a.get("esg_questions", {})
        if not q:
            continue
        tid = q.get("theme_id")
        qid = q.get("question_id")
        if tid and qid:
            theme_name = theme_map.get(tid, {}).get("name", "")
            answer_lookup[(theme_name, qid)] = a

    # Copy template to temp directory
    output_dir = os.path.join(tempfile.gettempdir(), "erm_exports")
    os.makedirs(output_dir, exist_ok=True)
    safe_name = re.sub(r'[^\w\s\-]', '', company["name"]).strip()
    filename = f"{safe_name} {analysis['report_year']}.xlsm"
    filepath = os.path.join(output_dir, filename)
    shutil.copy2(TEMPLATE_PATH, filepath)

    # Open the copy preserving VBA macros
    wb = load_workbook(filepath, keep_vba=True)

    # Resolve GICS sector for Materialidade formula compatibility
    gics_sector = _normalize_sector_to_gics(company.get("sector", ""))

    # --- Fill CAPA sheet ---
    if "Capa" in wb.sheetnames:
        ws_capa = wb["Capa"]
        for row_num, col_num in _find_company_name_cells(ws_capa):
            ws_capa.cell(row=row_num, column=col_num).value = company["name"]
        for row_num, col_num in _find_sector_cells(ws_capa):
            ws_capa.cell(row=row_num, column=col_num).value = gics_sector
        # Explicitly set G35 (critical: Peso formulas reference Capa!$G$35)
        ws_capa["G35"] = gics_sector
        # Fill date (G34) and analyst (G33)
        ws_capa["G34"] = datetime.now().strftime("%d/%m/%Y")
        ws_capa["G33"] = "Analise Automatizada (IA)"

    # --- Fill 'Dados da empresa' sheet ---
    if "Dados da empresa" in wb.sheetnames:
        ws_dados = wb["Dados da empresa"]
        ws_dados["C4"] = company["name"]
        ws_dados["C5"] = gics_sector
        ws_dados["C7"] = company.get("ticker", "")

    # --- Fill company name in ALL theme sheets + Resultado ---
    for sheet_name in wb.sheetnames:
        if sheet_name in [
            "Capa", "Sumário", "Conceito&Orientações", "Dados da empresa",
            "Abrev_Temas", "Assinaturas", "Padronização_dados", "Padronização",
            "Notas",
        ]:
            continue
        ws = wb[sheet_name]
        for row_num, col_num in _find_company_name_cells(ws):
            ws.cell(row=row_num, column=col_num).value = company["name"]
        for row_num, col_num in _find_sector_cells(ws):
            ws.cell(row=row_num, column=col_num).value = gics_sector

    # --- Fill answers in each theme sheet ---
    filled_count = 0
    for theme in themes:
        theme_name = theme["name"]
        sheet_name = THEME_TO_SHEET.get(theme_name)
        if not sheet_name or sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        qid_to_row = _build_question_row_map(ws)

        for qid, row_num in qid_to_row.items():
            answer_data = answer_lookup.get((theme_name, qid))
            if not answer_data:
                continue

            # Column E = Answer (Sim/Não/N/A)
            ans_val = _sanitize_for_excel(answer_data.get("answer"))
            if ans_val:
                ws.cell(row=row_num, column=5).value = ans_val

            # Column F = Justification
            justification = _sanitize_for_excel(answer_data.get("justification"))
            if justification:
                ws.cell(row=row_num, column=6).value = justification

            # Column L = Source reference
            source = _sanitize_for_excel(answer_data.get("source_reference"))
            if source:
                ws.cell(row=row_num, column=12).value = source

            # Column P = Improvement points
            improvements = _sanitize_for_excel(answer_data.get("improvement_points"))
            if improvements:
                ws.cell(row=row_num, column=16).value = improvements

            filled_count += 1

    wb.save(filepath)
    wb.close()

    print(f"[ExcelExporter] Filled {filled_count} answers into template -> {filepath}")
    return filepath
